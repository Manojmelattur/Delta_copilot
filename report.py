# =============================================================================
# report.py - Backtest Report Generator
# Handles: enhanced console table, CSV, Excel, HTML exports
# =============================================================================

import os
from datetime import datetime

import numpy as np
import pandas as pd

REPORT_DIR = "logs/reports"


# =============================================================================
# Metric Calculations
# =============================================================================


def _calc_profit_factor(trades: list) -> float:
    gross_wins = sum(t["net_pnl"] for t in trades if t["net_pnl"] > 0)
    gross_losses = abs(sum(t["net_pnl"] for t in trades if t["net_pnl"] <= 0))
    if gross_losses == 0:
        return float("inf")
    return round(gross_wins / gross_losses, 3)


def _calc_sharpe_ratio(trades: list) -> float:
    if len(trades) < 2:
        return 0.0
    pnls = [t["net_pnl"] for t in trades]
    mean = np.mean(pnls)
    std = np.std(pnls, ddof=1)
    if std == 0:
        return 0.0
    return round((mean / std) * np.sqrt(len(pnls)), 3)


def _calc_avg_duration(trades: list) -> str:
    if not trades:
        return "N/A"
    durations = []
    for t in trades:
        try:
            entry = pd.Timestamp(t["entry_time"])
            exit_ = pd.Timestamp(t["exit_time"])
            durations.append((exit_ - entry).total_seconds())
        except Exception:
            continue
    if not durations:
        return "N/A"
    avg_seconds = np.mean(durations)
    hours = int(avg_seconds // 3600)
    minutes = int((avg_seconds % 3600) // 60)
    return f"{hours}h {minutes}m"


def _calc_exit_breakdown(trades: list) -> dict:
    all_reasons = sorted(set(t["exit_reason"] for t in trades))
    breakdown = {}
    for reason in all_reasons:
        reason_trades = [t for t in trades if t["exit_reason"] == reason]
        reason_wins = [t for t in reason_trades if t["net_pnl"] > 0]
        reason_losses = [t for t in reason_trades if t["net_pnl"] <= 0]
        reason_pnl = sum(t["net_pnl"] for t in reason_trades)
        breakdown[reason] = {
            "count": len(reason_trades),
            "wins": len(reason_wins),
            "losses": len(reason_losses),
            "total_pnl": round(reason_pnl, 6),
            "avg_win": round(np.mean([t["net_pnl"] for t in reason_wins]), 6)
            if reason_wins
            else 0,
            "avg_loss": round(np.mean([t["net_pnl"] for t in reason_losses]), 6)
            if reason_losses
            else 0,
        }
    return breakdown


import re


def _get_strategy_name(strategy) -> str:
    """
    Safely resolve strategy name whether it is a @property or plain method.
    Extracts clean name from bound method string if needed.
    """
    name = strategy.name

    # If it is a callable (missing @property), call it first
    if callable(name):
        name = name()

    name = str(name)

    # If still looks like a bound method string, extract from it
    # e.g. "<bound method OrderBlockStrategy.name of <...>>"
    if name.startswith("<bound method"):
        match = re.search(r"<bound method (\w+)\.name", name)
        if match:
            class_name = match.group(1)  # e.g. "OrderBlockStrategy"
            # Strip "Strategy" suffix and split on camel case
            class_name = re.sub(r"Strategy$", "", class_name)
            # CamelCase -> "Camel Case"
            name = re.sub(r"(?<!^)(?=[A-Z])", " ", class_name).strip()

    return name


def build_strategy_metrics(name, trades: list, summary: dict) -> dict:
    clean_name = str(name)

    if not trades or not summary:
        return {
            "strategy": clean_name,
            "total_trades": 0,
            "wins": 0,
            "losses": 0,
            "win_rate": 0.0,
            "loss_rate": 0.0,  # <-- new
            "avg_win_loss_ratio": 0.0,  # <-- new
            "total_pnl_btc": 0.0,
            "avg_win_btc": 0.0,
            "avg_loss_btc": 0.0,
            "max_drawdown_btc": 0.0,
            "profit_factor": 0.0,
            "sharpe_ratio": 0.0,
            "avg_duration": "N/A",
            "exit_breakdown": {},
        }

    total = summary.get("total_trades", 0)
    losses = summary.get("losses", 0)
    avg_win = summary.get("avg_win_btc", 0.0)
    avg_loss = summary.get("avg_loss_btc", 0.0)

    loss_rate = round((losses / total * 100), 1) if total > 0 else 0.0

    # Avg Win / Avg Loss ratio — how much you win vs how much you lose per trade
    if avg_loss != 0:
        avg_win_loss_ratio = round(abs(avg_win / avg_loss), 3)
    else:
        avg_win_loss_ratio = float("inf")

    return {
        "strategy": clean_name,
        "total_trades": total,
        "wins": summary.get("wins", 0),
        "losses": losses,
        "win_rate": summary.get("win_rate", 0.0),
        "loss_rate": loss_rate,  # <-- new
        "avg_win_loss_ratio": avg_win_loss_ratio,  # <-- new
        "total_pnl_btc": summary.get("total_pnl_btc", 0.0),
        "avg_win_btc": avg_win,
        "avg_loss_btc": avg_loss,
        "max_drawdown_btc": summary.get("max_drawdown_btc", 0.0),
        "profit_factor": _calc_profit_factor(trades),
        "sharpe_ratio": _calc_sharpe_ratio(trades),
        "avg_duration": _calc_avg_duration(trades),
        "exit_breakdown": _calc_exit_breakdown(trades),
    }


# =============================================================================
# Console Table
# =============================================================================


def print_comparison_table(results: dict) -> list:
    all_metrics = []
    for name, (trades, summary) in results.items():
        all_metrics.append(build_strategy_metrics(name, trades, summary))

    # Collect all unique exit reasons
    all_reasons = sorted(
        set(reason for m in all_metrics for reason in m["exit_breakdown"].keys())
    )

    # Column widths
    S = 15  # Strategy
    T = 6  # Trades
    W = 5  # Wins
    L = 6  # Losses
    WR = 7  # Win%
    LR = 7  # Loss%        <-- new
    WLR = 8  # W/L Ratio    <-- new
    P = 11  # PnL
    DD = 10  # MaxDD
    PF = 7  # PF
    SH = 8  # Sharpe
    AD = 9  # AvgDur
    AW = 11  # AvgWin
    AL = 11  # AvgLoss
    ER = 9  # Exit reason col width

    fixed_width = S + T + W + L + WR + LR + WLR + P + DD + PF + SH + AD + AW + AL + 32
    exit_width = len(all_reasons) * (ER + 1)
    total_width = fixed_width + exit_width

    # Header
    print(f"\n{'=' * total_width}")
    print(f"{'  MULTI-STRATEGY BACKTEST COMPARISON':^{total_width}}")
    print(f"{'=' * total_width}")

    # Group label row
    exit_group = "".join(f"{r:^{ER + 1}}" for r in all_reasons)
    print(
        f"  {'':^{S}} {'':^{T}} {'W/L':^{W + L + 2}}"
        f" {'WIN%':^{WR}} {'LOSS%':^{LR}} {'W/L':^{WLR}}"
        f" {'':^{P}} {'':^{DD}} {'':^{PF}} {'':^{SH}}"
        f" {'':^{AD}} {'--- AVG ---':^{AW + AL + 2}}"
        f" {'--- EXIT REASONS (count: W/L) ---':^{exit_width}}"
    )

    # Column name row
    exit_headers = "".join(f"{r[:ER]:^{ER + 1}}" for r in all_reasons)
    print(
        f"  {'Strategy':<{S}} {'Trades':>{T}} {'Wins':>{W}} {'Losses':>{L}}"
        f" {'Win%':>{WR}} {'Loss%':>{LR}} {'Ratio':>{WLR}}"
        f" {'PnL(BTC)':>{P}} {'MaxDD':>{DD}}"
        f" {'PF':>{PF}} {'Sharpe':>{SH}} {'AvgDur':>{AD}}"
        f" {'AvgWin':>{AW}} {'AvgLoss':>{AL}}"
        f" {exit_headers}"
    )
    print(f"  {'-' * (total_width - 2)}")

    # Data rows
    for m in all_metrics:
        pnl_str = f"{m['total_pnl_btc']:+.5f}"
        dd_str = f"{m['max_drawdown_btc']:.5f}"
        pf_str = (
            f"{m['profit_factor']:.3f}" if m["profit_factor"] != float("inf") else "inf"
        )
        sharpe_str = f"{m['sharpe_ratio']:+.3f}"
        aw_str = f"{m['avg_win_btc']:+.5f}"
        al_str = f"{m['avg_loss_btc']:+.5f}"
        lr_str = f"{m['loss_rate']:.1f}%"
        wlr_str = (
            f"{m['avg_win_loss_ratio']:.3f}"
            if m["avg_win_loss_ratio"] != float("inf")
            else "inf"
        )

        # Exit reason columns
        exit_cols = ""
        for reason in all_reasons:
            stats = m["exit_breakdown"].get(reason)
            if stats:
                cell = f"{stats['count']}:{stats['wins']}/{stats['losses']}"
            else:
                cell = "-"
            exit_cols += f"{cell:^{ER + 1}}"

        print(
            f"  {m['strategy']:<{S}}"
            f" {m['total_trades']:>{T}}"
            f" {m['wins']:>{W}}"
            f" {m['losses']:>{L}}"
            f" {m['win_rate']:>{WR - 1}.1f}%"
            f" {lr_str:>{LR}}"
            f" {wlr_str:>{WLR}}"
            f" {pnl_str:>{P}}"
            f" {dd_str:>{DD}}"
            f" {pf_str:>{PF}}"
            f" {sharpe_str:>{SH}}"
            f" {m['avg_duration']:>{AD}}"
            f" {aw_str:>{AW}}"
            f" {al_str:>{AL}}"
            f" {exit_cols}"
        )

    # Footer totals row
    print(f"  {'-' * (total_width - 2)}")
    total_trades = sum(m["total_trades"] for m in all_metrics)
    total_wins = sum(m["wins"] for m in all_metrics)
    total_losses = sum(m["losses"] for m in all_metrics)
    total_pnl = sum(m["total_pnl_btc"] for m in all_metrics)
    avg_wr = (
        sum(m["win_rate"] for m in all_metrics) / len(all_metrics)
        if all_metrics
        else 0.0
    )
    avg_lr = (
        sum(m["loss_rate"] for m in all_metrics) / len(all_metrics)
        if all_metrics
        else 0.0
    )
    avg_wlr = (
        sum(
            m["avg_win_loss_ratio"]
            for m in all_metrics
            if m["avg_win_loss_ratio"] != float("inf")
        )
        / len([m for m in all_metrics if m["avg_win_loss_ratio"] != float("inf")])
        if any(m["avg_win_loss_ratio"] != float("inf") for m in all_metrics)
        else 0.0
    )

    footer_exit = ""
    for reason in all_reasons:
        total_cnt = sum(
            m["exit_breakdown"][reason]["count"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        total_w = sum(
            m["exit_breakdown"][reason]["wins"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        total_l = sum(
            m["exit_breakdown"][reason]["losses"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        footer_exit += f"{f'{total_cnt}:{total_w}/{total_l}':^{ER + 1}}"

    print(
        f"  {'TOTAL':<{S}}"
        f" {total_trades:>{T}}"
        f" {total_wins:>{W}}"
        f" {total_losses:>{L}}"
        f" {avg_wr:>{WR - 1}.1f}%"
        f" {avg_lr:>{LR - 1}.1f}%"
        f" {avg_wlr:>{WLR}.3f}"
        f" {total_pnl:>+{P}.5f}"
        f" {'':>{DD}}"
        f" {'':>{PF}}"
        f" {'':>{SH}}"
        f" {'':>{AD}}"
        f" {'':>{AW}}"
        f" {'':>{AL}}"
        f" {footer_exit}"
    )
    print(f"{'=' * total_width}\n")

    return all_metrics

    # ------------------------------------------------------------------
    # Footer totals row
    # ------------------------------------------------------------------
    print(f"  {'-' * (total_width - 2)}")
    total_trades = sum(m["total_trades"] for m in all_metrics)
    total_wins = sum(m["wins"] for m in all_metrics)
    total_losses = sum(m["losses"] for m in all_metrics)
    total_pnl = sum(m["total_pnl_btc"] for m in all_metrics)
    avg_wr = (
        sum(m["win_rate"] for m in all_metrics) / len(all_metrics)
        if all_metrics
        else 0.0
    )

    # Footer exit reason totals
    footer_exit = ""
    for reason in all_reasons:
        total_cnt = sum(
            m["exit_breakdown"][reason]["count"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        total_w = sum(
            m["exit_breakdown"][reason]["wins"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        total_l = sum(
            m["exit_breakdown"][reason]["losses"]
            for m in all_metrics
            if reason in m["exit_breakdown"]
        )
        cell = f"{total_cnt}:{total_w}/{total_l}"
        footer_exit += f"{cell:^{ER + 1}}"

    print(
        f"  {'TOTAL':<{S}}"
        f" {total_trades:>{T}}"
        f" {total_wins:>{W}}"
        f" {total_losses:>{L}}"
        f" {avg_wr:>{WR - 1}.1f}%"
        f" {total_pnl:>+{P}.5f}"
        f" {'':>{DD}}"
        f" {'':>{PF}}"
        f" {'':>{SH}}"
        f" {'':>{AD}}"
        f" {'':>{AW}}"
        f" {'':>{AL}}"
        f" {footer_exit}"
    )
    print(f"{'=' * total_width}\n")

    return all_metrics


# =============================================================================
# CSV Export
# =============================================================================


def export_csv(all_metrics: list, symbol: str, timeframe: str) -> str:
    """Export flat comparison table to CSV."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        REPORT_DIR, f"backtest_{symbol}_{timeframe}_{timestamp}.csv"
    )

    rows = []
    for m in all_metrics:
        row = {k: v for k, v in m.items() if k != "exit_breakdown"}
        # Flatten exit breakdown into columns
        for reason, stats in m["exit_breakdown"].items():
            row[f"exit_{reason}_count"] = stats["count"]
            row[f"exit_{reason}_wins"] = stats["wins"]
            row[f"exit_{reason}_losses"] = stats["losses"]
            row[f"exit_{reason}_pnl"] = stats["total_pnl"]
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_csv(filepath, index=False)
    print(f"[REPORT] CSV saved to: {filepath}")
    return filepath


# =============================================================================
# Excel Export
# =============================================================================


def export_excel(all_metrics: list, symbol: str, timeframe: str) -> str:
    """
    Export multi-sheet Excel report:
      Sheet 1: Summary comparison table
      Sheet 2: Exit reason breakdown (all strategies)
      Sheet 3+: Per-strategy exit breakdown
    """
    os.makedirs(REPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        REPORT_DIR, f"backtest_{symbol}_{timeframe}_{timestamp}.xlsx"
    )

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("[REPORT] openpyxl not installed. Run: pip install openpyxl")
        return ""

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_cols = [
        "strategy",
        "total_trades",
        "wins",
        "losses",
        "win_rate",
        "total_pnl_btc",
        "avg_win_btc",
        "avg_loss_btc",
        "max_drawdown_btc",
        "profit_factor",
        "sharpe_ratio",
        "avg_duration",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary.append(summary_cols)
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in all_metrics:
        row = [m.get(c, "") for c in summary_cols]
        ws_summary.append(row)

    # Color PnL column: green positive, red negative
    pnl_col_idx = summary_cols.index("total_pnl_btc") + 1
    for row_idx, m in enumerate(all_metrics, start=2):
        cell = ws_summary.cell(row=row_idx, column=pnl_col_idx)
        if isinstance(cell.value, float):
            cell.font = Font(color="00B050" if cell.value >= 0 else "FF0000", bold=True)

    # Auto-width
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_summary.column_dimensions[col[0].column_letter].width = max_len + 4

    # ------------------------------------------------------------------
    # Sheet 2: Exit Breakdown (all strategies combined)
    # ------------------------------------------------------------------
    ws_exit = wb.create_sheet("Exit Breakdown")
    exit_headers = [
        "strategy",
        "exit_reason",
        "count",
        "wins",
        "losses",
        "total_pnl",
        "avg_win",
        "avg_loss",
    ]
    ws_exit.append(exit_headers)
    for cell in ws_exit[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in all_metrics:
        for reason, stats in m["exit_breakdown"].items():
            ws_exit.append(
                [
                    m["strategy"],
                    reason,
                    stats["count"],
                    stats["wins"],
                    stats["losses"],
                    stats["total_pnl"],
                    stats["avg_win"],
                    stats["avg_loss"],
                ]
            )

    for col in ws_exit.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_exit.column_dimensions[col[0].column_letter].width = max_len + 4

    # ------------------------------------------------------------------
    # Sheet 3+: Per-strategy exit breakdown
    # ------------------------------------------------------------------
    for m in all_metrics:
        if not m["exit_breakdown"]:
            continue
        sheet_name = str(m["strategy"])[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        ws.append(
            [
                "exit_reason",
                "count",
                "wins",
                "losses",
                "total_pnl",
                "avg_win",
                "avg_loss",
            ]
        )
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for reason, stats in m["exit_breakdown"].items():
            ws.append(
                [
                    reason,
                    stats["count"],
                    stats["wins"],
                    stats["losses"],
                    stats["total_pnl"],
                    stats["avg_win"],
                    stats["avg_loss"],
                ]
            )
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filepath)
    print(f"[REPORT] Excel saved to: {filepath}")
    return filepath


# =============================================================================
# HTML Export
# =============================================================================


def export_html(all_metrics: list, symbol: str, timeframe: str) -> str:
    """Export a styled HTML report with summary table and exit breakdowns."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        REPORT_DIR, f"backtest_{symbol}_{timeframe}_{timestamp}.html"
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def pnl_color(val):
        try:
            return (
                "color:#00B050;font-weight:bold"
                if float(val) >= 0
                else "color:#FF0000;font-weight:bold"
            )
        except Exception:
            return ""

    rows_html = ""
    for m in all_metrics:
        pf = (
            f"{m['profit_factor']:.3f}" if m["profit_factor"] != float("inf") else "inf"
        )
        rows_html += f"""
        <tr>
            <td>{m["strategy"]}</td>
            <td>{m["total_trades"]}</td>
            <td>{m["wins"]}</td>
            <td>{m["losses"]}</td>
            <td>{m["win_rate"]:.1f}%</td>
            <td style="{pnl_color(m["total_pnl_btc"])}">{m["total_pnl_btc"]:+.6f}</td>
            <td style="{pnl_color(m["avg_win_btc"])}">{m["avg_win_btc"]:+.6f}</td>
            <td style="{pnl_color(m["avg_loss_btc"])}">{m["avg_loss_btc"]:+.6f}</td>
            <td>{m["max_drawdown_btc"]:.6f}</td>
            <td>{pf}</td>
            <td style="{pnl_color(m["sharpe_ratio"])}">{m["sharpe_ratio"]:+.3f}</td>
            <td>{m["avg_duration"]}</td>
        </tr>"""

    # --- Exit breakdown tables ---
    exit_tables_html = ""
    for m in all_metrics:
        if not m["exit_breakdown"]:
            continue
        exit_rows = ""
        for reason, stats in m["exit_breakdown"].items():
            avg_win_str = f"{stats['avg_win']:+.4f}" if stats["wins"] > 0 else "N/A"
            avg_loss_str = f"{stats['avg_loss']:+.4f}" if stats["losses"] > 0 else "N/A"
            exit_rows += f"""
            <tr>
                <td>{reason}</td>
                <td>{stats["count"]}</td>
                <td>{stats["wins"]}</td>
                <td>{stats["losses"]}</td>
                <td style="{pnl_color(stats["total_pnl"])}">{stats["total_pnl"]:+.4f}</td>
                <td style="{pnl_color(stats["avg_win"]) if stats["wins"] > 0 else ""}">{avg_win_str}</td>
                <td style="{pnl_color(stats["avg_loss"]) if stats["losses"] > 0 else ""}">{avg_loss_str}</td>
            </tr>"""

        exit_tables_html += f"""
        <h3 style="color:#1F4E79;margin-top:30px">{m["strategy"]} — Exit Reason Breakdown</h3>
        <table>
            <thead>
                <tr>
                    <th>Reason</th><th>Count</th><th>Wins</th><th>Losses</th>
                    <th>Total PnL</th><th>Avg Win</th><th>Avg Loss</th>
                </tr>
            </thead>
            <tbody>{exit_rows}</tbody>
        </table>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Backtest Report - {symbol} {timeframe}</title>
    <style>
        body {{ font-family: Arial, sans-serif; background: #f4f6f9; color: #222; margin: 30px; }}
        h1 {{ color: #1F4E79; }}
        h2 {{ color: #2E75B6; margin-top: 40px; }}
        h3 {{ color: #1F4E79; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 10px; background: #fff; }}
        th {{ background: #1F4E79; color: #fff; padding: 8px 12px; text-align: center; font-size: 13px; }}
        td {{ padding: 7px 12px; border-bottom: 1px solid #ddd; text-align: center; font-size: 13px; }}
        tr:hover {{ background: #eaf1fb; }}
        .meta {{ color: #555; font-size: 13px; margin-bottom: 20px; }}
        .footer {{ margin-top: 40px; color: #999; font-size: 12px; }}
    </style>
</head>
<body>
    <h1>Backtest Report</h1>
    <div class="meta">
        <strong>Symbol:</strong> {symbol} &nbsp;|&nbsp;
        <strong>Timeframe:</strong> {timeframe} &nbsp;|&nbsp;
        <strong>Generated:</strong> {generated_at}
    </div>

    <h2>Strategy Comparison</h2>
    <table>
        <thead>
            <tr>
                <th>Strategy</th><th>Trades</th><th>Wins</th><th>Losses</th>
                <th>Win%</th><th>Total PnL (BTC)</th><th>Avg Win</th><th>Avg Loss</th>
                <th>Max DD</th><th>Profit Factor</th><th>Sharpe</th><th>Avg Duration</th>
            </tr>
        </thead>
        <tbody>{rows_html}</tbody>
    </table>

    <h2>Exit Reason Breakdown</h2>
    {exit_tables_html}

    <div class="footer">Generated by Delta Copilot Backtester &mdash; {generated_at}</div>
</body>
</html>"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"[REPORT] HTML saved to: {filepath}")
    return filepath


# =============================================================================
# Master Export — call this from backtester.py
# =============================================================================


def generate_all_reports(
    results: dict,
    symbol: str = "BTCUSD",
    timeframe: str = "15m",
) -> list:
    """
    Entry point called from backtester.py run_all_backtests().
    Prints console table and exports CSV, Excel, HTML.
    Returns list of metrics dicts.
    """
    all_metrics = print_comparison_table(results)
    export_csv(all_metrics, symbol, timeframe)
    export_excel(all_metrics, symbol, timeframe)
    export_html(all_metrics, symbol, timeframe)
    return all_metrics
